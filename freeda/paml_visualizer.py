#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun May  9 14:12:04 2021

@author: damian
"""

from freeda import fasta_reader
from Bio import AlignIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from Bio import pairwise2
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from json import dump
import re
import os
import shutil
import logging
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import use
import pandas as pd
import math


def analyse_PAML_results(wdir, result_path, all_genes, nr_of_species_total_dict,
                         ref_species, PAML_logfile_name, day, genes_under_positive_selection, failed_paml):
    """Analyses PAML results for each gene unless no PAML result available"""

    all_matched_adaptive_sites_ref = {}
    
    for gene in all_genes:

        if gene not in failed_paml:
        
            gene_folder = result_path + gene + "/"
            if os.path.exists(gene_folder + gene + "_PAML_analysis.tif"):
                message = "\n***************\n\n PAML analysis graph for : %s already exists" % gene
                print(message)
                logging.info(message)
        
            else:
                nr_of_species_total = nr_of_species_total_dict[gene]
                matched_adaptive_sites_ref = plot_PAML(wdir, result_path, gene,
                                                   nr_of_species_total, ref_species,
                                                   genes_under_positive_selection)
                all_matched_adaptive_sites_ref[gene] = matched_adaptive_sites_ref

    # get protein alignment that matches 3D structure
    for gene in all_genes:
        try:
            get_alignment_matching_structure(result_path, ref_species, gene, all_matched_adaptive_sites_ref)
        # if all_matched_adaptive_sites_ref dictionary was not generated (all genes failed paml)
        except KeyError:
            return

    # prepare a dict with PAML stats
    final_PAML_log_dict = read_output_PAML(result_path, PAML_logfile_name, all_matched_adaptive_sites_ref, failed_paml)
    # generate a PAML result excel sheet
    output_excel_sheet(wdir, final_PAML_log_dict, result_path, day)
    
    # save the matched adaptive sites into a file
    with open("all_matched_adaptive_sites_ref.txt", "w") as file:
        dump(all_matched_adaptive_sites_ref, file)
    
    shutil.move(wdir + "all_matched_adaptive_sites_ref.txt", result_path + "all_matched_adaptive_sites_ref.txt")

    return final_PAML_log_dict


def get_alignment_matching_structure(result_path, ref_species, gene, dictionary):
    """Generates gene alignment matching 3D structure"""

    # get gene alignment corresponding to unedited reference gene
    aa_features_dict = dictionary[gene]
    filename = gene + "_protein_alignment.fasta"
    all_seq_dict = fasta_reader.alignment_file_to_dict(result_path, ref_species, filename)

    # remove alignment file (not needed)
    os.remove(result_path + filename)

    # make temporary dictionary donor of aa
    temp_seq_dict = {}
    for species in all_seq_dict:
        temp_seq_dict[species] = []
        for position, aa in enumerate(all_seq_dict[species]):
            temp_seq_dict[species].append(aa)

    # make a new dictionary to hold empty positions ("-") or aa from temp_seq_dict
    new_seq_dict = {}
    for species in all_seq_dict:
        new_seq_dict[species] = {}
        for position, aa in enumerate(aa_features_dict):
            new_seq_dict[species][position] = ""

    # fill up new_seq_dict
    for species in all_seq_dict:

        for position, features in aa_features_dict.items():

            # reference species sequence is full length
            if species.lstrip(">") == ref_species:
                new_seq_dict[species][position] = features[0]
                continue

            position = int(position) - 1
            # add aa
            if features[-1] == 1:
                new_seq_dict[species][position] = temp_seq_dict[species].pop(0)
            # introduce a dash (missing aa)
            else:
                new_seq_dict[species][position] = "-"

    with open(result_path + gene + "_protein_alignment.fasta", "w") as f:
        for species in new_seq_dict:
            f.write(species + "\n")
            # reconstruct sequence
            seq = "".join([aa for position, aa in new_seq_dict[species].items()])
            f.write(seq + "\n")


def read_output_PAML(result_path, PAML_logfile_name, all_matched_adaptive_sites_ref, failed_paml):
    """Reads the output PAML and finds LRTs and computes p-values for M2a-M1a and M8-M7 models"""

    # prepare the dict storing the PAML result
    PAML_log_dict = {"Gene name": [],
                        "Nr of species analyzed": [],
                        "Species": [],
                        "CDS Coverage": [],
                        "M2a vs M1a (LRT)": [],
                        "M2a vs M1a (p-value)": [],
                        "M8 vs M7 (LRT)": [],
                        "M8 vs M7 (p-value)": [],
                        "Sites with pr < 0.90": [],
                        "Sites with pr >= 0.90": []}

    # mark genes that failed paml
    if failed_paml:
        for gene in failed_paml:
            for key, values in PAML_log_dict.items():
                if key == "Gene name":
                    PAML_log_dict[key].append(gene)
                else:
                    PAML_log_dict[key].append("-")

    # extract PAML results from PAML log file
    PAML_log_path = result_path + PAML_logfile_name
    
    with open(PAML_log_path, "r") as f:
        file = f.readlines()
        
        start_recording = False

        for line in file:
            
            # record gene name
            if line.startswith(" --------- *"):
                gene = line.replace("-", "").replace(" ", "").replace("*", "").rstrip("\n")
                if gene in failed_paml:
                    continue
                PAML_log_dict["Gene name"].append(gene)
                start_recording = True 

            # record species
            if start_recording is True and line.startswith(" Final species"):
                species = line.split(":")[1].split("[")[1].replace("]", "").replace("\n", "").replace("'", "")
                nr_species = line.split(":")[1].split(" ")[1]
                PAML_log_dict["Nr of species analyzed"].append(nr_species)
                PAML_log_dict["Species"].append(species)
                
            # record CDS_coverage
            if start_recording is True and line.startswith("['Original alignment:"):
                coverage = line.split(")")[0].split("(")[-1].replace(" ", "")
                PAML_log_dict["CDS Coverage"].append(coverage)
            
            # record LRTs
            if start_recording is True and line.startswith(" PAML LRTs"):
                
                LRT1 = line.split(":")[1].split("and")[0].split("-")
                if LRT1[1].endswith("e") is True:
                    M2a_vs_M1a_LRT = LRT1[1].replace(" ", "") + "-" + LRT1[2]
                if LRT1[1].endswith("e") is False:
                    M2a_vs_M1a_LRT = LRT1[1].replace(" ", "")
                if M2a_vs_M1a_LRT != "None":
                    M2a_vs_M1a_LRT = round(float(M2a_vs_M1a_LRT), 4)
                    if M2a_vs_M1a_LRT == 0.0:
                        M2a_vs_M1a_LRT = 0.0001

                LRT2 = line.split(":")[1].split("and")[1].split("-")
                # catch "e-" notation because it gets split
                if LRT2[1].endswith("e") is True:
                    M8_vs_M7_LRT = LRT2[1].replace(" ", "") + "-" + LRT2[2].rstrip("\n")
                if LRT2[1].endswith("e") is False:
                    M8_vs_M7_LRT = LRT2[1].replace(" ", "").rstrip("\n")
                if M8_vs_M7_LRT != "None":
                    M8_vs_M7_LRT = round(float(M8_vs_M7_LRT), 4)
                    if M8_vs_M7_LRT == 0.0:
                        M8_vs_M7_LRT = 0.0001
                    
                PAML_log_dict["M2a vs M1a (LRT)"].append(M2a_vs_M1a_LRT)
                PAML_log_dict["M8 vs M7 (LRT)"].append(M8_vs_M7_LRT)
            
            # record p_values
            if start_recording is True and line.startswith(" -> PAML p-values"):
                
                p_value1 = line.split(":")[1].split("and")[0].split("-")
                # catch "e-" notation because it gets split
                if p_value1[1].endswith("e") is True:
                    M2a_vs_M1a_pvalue = p_value1[1].replace(" ","") + "-" + p_value1[2]
                if p_value1[1].endswith("e") is False:
                    M2a_vs_M1a_pvalue = p_value1[1].replace(" ","")
                if M2a_vs_M1a_pvalue != "None":
                    M2a_vs_M1a_pvalue = round(float(M2a_vs_M1a_pvalue), 4)
                    if M2a_vs_M1a_pvalue == 0.0:
                        M2a_vs_M1a_pvalue = 0.0001
                    
                PAML_log_dict["M2a vs M1a (p-value)"].append(M2a_vs_M1a_pvalue)
                
                p_value2 = line.split(":")[1].split("and")[1].split("-")
                # catch "e-" notation because it gets split
                if p_value2[1].endswith("e"):
                    M8_vs_M7_pvalue = p_value2[1].replace(" ","") + "-" + p_value2[2].rstrip("\n")
                if p_value2[1].endswith("e") is False:
                    M8_vs_M7_pvalue = p_value2[1].replace(" ","").rstrip("\n")
                if M8_vs_M7_pvalue != "None":
                    M8_vs_M7_pvalue = round(float(M8_vs_M7_pvalue), 4)
                    if M8_vs_M7_pvalue == 0.0:
                        M8_vs_M7_pvalue = 0.0001

                PAML_log_dict["M8 vs M7 (p-value)"].append(M8_vs_M7_pvalue)
                    
                # does not report adaptive sites in genes failing M8 vs M7 test
                if M8_vs_M7_pvalue == "None" or float(M8_vs_M7_pvalue) >= 0.05:
                    PAML_log_dict["Sites with pr < 0.90"].append("Not likely")
                    PAML_log_dict["Sites with pr >= 0.90"].append("Not likely")
                
                # remove adaptive sites from genes that are not rapidly evolving
                if M8_vs_M7_pvalue != "None" and float(M8_vs_M7_pvalue) <= 0.05:
                        
                    matched_dict = all_matched_adaptive_sites_ref[gene]
                    mild_sites = {}
                    strong_sites = {}
        
                    for site, values in matched_dict.items():

                        if float(values[2]) > 0.00:
                            adaptive_site = values[0] + str(site)
                            probability = float(values[2])
                                
                            if float(values[2]) < 0.90:
                                mild_sites[adaptive_site] = probability
    
                            if float(values[2]) >= 0.90:
                                strong_sites[adaptive_site] = probability
                        
                    mild_sites_to_append = (str(len(mild_sites)) + " " + str(mild_sites)).replace("'","")
                    strong_sites_to_append = (str(len(strong_sites)) + " " + str(strong_sites)).replace("'","")
                        
                    PAML_log_dict["Sites with pr < 0.90"].append(mild_sites_to_append)
                    PAML_log_dict["Sites with pr >= 0.90"].append(strong_sites_to_append)

                # finish recording loop through the log file
                start_recording = False
                
        final_PAML_log_dict = PAML_log_dict
        
    return final_PAML_log_dict


def output_excel_sheet(wdir, final_PAML_log_dict, result_path, day):
    """Makes an excel sheet with most important PAML results"""

    # make an empty excel sheet using openpyxl library
    wb = Workbook()
    ws = wb.active
    
    # set column width
    columns_to_format = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    for c in columns_to_format:
        ws.column_dimensions[c].width = 19

    # format headers (cell by cell)
    headers = [ws["A1"], ws["B1"], ws["C1"], ws["D1"], ws["E1"], 
               ws["F1"], ws["G1"], ws["H1"], ws["I1"], ws["J1"]]
    heads = list(final_PAML_log_dict.keys())
    header_count = 1
    for header in headers:
        
        # THIS IS MOST LIKELY NOT NEEDED (CHECK)
        head = heads.pop(0)
        ws.cell(row=1, column=header_count, value=head)
        
        header.font = Font(bold = True, name = "Arial", size = 10)
        double = Side(border_style = "thin")
        header.border = Border(top = double, left = double, right = double, bottom = double)
        header.alignment = Alignment(horizontal = "center", vertical = "center")
        header.fill = PatternFill("solid", fgColor = "00FFCC00")    
        header_count += 1
    
    all_rows = []
    for entry in final_PAML_log_dict.values():
    
        # make it a dataframe
        df = pd.DataFrame(entry)

        # read in all values from each key of final_PAML_log_dict (remove index column)
        rows = dataframe_to_rows(df, index=False)
        all_rows.append(list(rows))
    
    # counters for the iteration over cells in the worksheet (openpyxl index starts at 1)
    # leave the first row for headers
    column_count = 1
    row_count = 2
    
    # go over each entry per list of entries from the final PAML log dict
    for row in all_rows:

        for one_entry in row:
            
            # need to convert the generator into a string to tidy it
            entry_as_string = str(one_entry).replace("[", "").replace("]", "").replace("'", "").replace("\n", "")
            one_entry = entry_as_string
            
            # reset the row counter
            if row_count == len(list(row))+1:
                column_count += 1
                row_count = 2
            
            # record a header of new column
            if one_entry == "0":
                head = [key for number, key in enumerate(final_PAML_log_dict.keys(),1) if number == column_count][0]
                
                # always goes into the first row
                ws.cell(row=1, column=column_count, value=head.replace("'",""))
            
            # record entry
            else:
                ws.cell(row=row_count, column=column_count, value=one_entry)
                row_count += 1
    
    # make all rows centered
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal = "center", vertical = "center")
    
    # save the excel document
    excel_filename = "PAML_result" + day + ".xlsx"
    wb.save(excel_filename)
    shutil.move(wdir + excel_filename, result_path)


def plot_PAML(wdir, result_path, gene, nr_of_species_total, ref_species, genes_under_positive_selection):
    """Maps PAML result onto the cds of reference species and outputs it as a bar graph"""
    
    # get ref and final aa sequence for a gene
    ref_sequence_record, final_sequence_record = get_ref_and_final_seqs(wdir, gene, result_path, ref_species)
    # organise them into easy to search dictionaries
    ref_species_dict, final_ref_species_dict, final_length = organise_ref_and_final_seqs(ref_sequence_record, final_sequence_record)
    # map the aa residues between the ref and final sequences
    mapped_ref_and_final_residues_dict = map_ref_and_final_residues(ref_sequence_record, final_sequence_record)
    # find dN/dS omega ratio per site based on "rst" file (final only for now)
    omega_dict = get_omegas(gene, result_path, final_length)
    # read PAML output file to find which residues are rapidly evolving in the final seq
    adaptive_sites_dict = get_adaptive_sites(result_path, gene)
    # find these residues in the ref sequence
    matched_adaptive_sites_ref = match_adaptive_sites_to_ref(final_ref_species_dict,
            mapped_ref_and_final_residues_dict, adaptive_sites_dict, omega_dict)
    # mark the sites that were not analyzed by PAML
    final_dict_to_plot = mark_skipped_sites(matched_adaptive_sites_ref, mapped_ref_and_final_residues_dict)
    # record and write breakdown of adaptive sites overlay to ref cds
    record_adaptive_sites(final_dict_to_plot, gene, genes_under_positive_selection)
    # plot omegas and probabilities
    make_graphs(wdir, ref_species, final_dict_to_plot, result_path, gene,
                nr_of_species_total, genes_under_positive_selection)

    return matched_adaptive_sites_ref


def mark_skipped_sites(matched_adaptive_sites_ref, mapped_ref_and_final_residues_dict):
    """Marks sites that were not analyzed by PAML"""

    for site in mapped_ref_and_final_residues_dict:
        if mapped_ref_and_final_residues_dict[site][0] != "-":
            matched_adaptive_sites_ref[site].append(1)
        else:
            matched_adaptive_sites_ref[site].append(0)
            
    return matched_adaptive_sites_ref


def record_adaptive_sites(final_dict_to_plot, gene, genes_under_positive_selection):
    """Makes a uniprot format representation of each residue in aa seq of reference species
    (number, selection, presence)"""

    annotate_selection = False
    if gene in genes_under_positive_selection:
        annotate_selection = True

    # row_residues and row_features are well set
    row_positions = ""
    row_residues = ""
    row_features = ""
    position_counter = 0
    row_counter = [row for row in range(1, len(final_dict_to_plot) + 1) if row % 10 == 0]
    
    for position, values in final_dict_to_plot.items():
        
        # break all rows for visibility
        if position_counter % 50 != 0 and position_counter != 0 and position_counter % 10 == 0:

            row_residues = row_residues + " "
            row_features = row_features + " "
        
        position_counter += 1
        
        # prepare a position count array
        # first row
        if position_counter == 1:
            position_marker = row_counter.pop(0)
            row_positions = 8 * " " + str(position_marker) + 9 * " "
        
        # any other row
        if (position_counter + 1) % 10 == 0 and row_counter != []:
            position_marker = row_counter.pop(0)
            
            if position_marker < 100:
                
                if position_marker % 50 == 0:
                    row_positions = row_positions + str(position_marker) + "\n" + 8 * " "
                else:
                    row_positions = row_positions + str(position_marker) + 9 * " "
                    
            if position_marker == 100:
                row_positions = row_positions[:-1] + str(position_marker) + "\n" + 7 * " "
                    
            if 1000 > position_marker > 100:
        
                if position_marker % 50 == 0:
                    row_positions = row_positions + str(position_marker) + "\n" + 7 * " "
                else:
                    row_positions = row_positions + str(position_marker) + 8 * " "
            
            if position_marker == 1000:
                row_positions = row_positions[:-1] + str(position_marker) + "\n" + 6 * " "
            
            if 1000 < position_marker:
            
                if position_marker % 50 == 0:
                    row_positions = row_positions + str(position_marker) + "\n" + 6 * " "
                else:
                    row_positions = row_positions + str(position_marker) + 7 * " "
        
        # Prepare residues and features arrays :

        # residue absent in PAML analysis
        if values[3] == 0:
            residue = values[0]
            row_features = row_features + "-"
            row_residues = row_residues + residue
            if position % 50 == 0:
                row_residues = row_residues + "\n"
                row_features = row_features + "\n"
            continue

        # do not annotate probabilities if gene not under positive selection
        if not annotate_selection:
            residue = values[0]
            row_features = row_features + " "
            row_residues = row_residues + residue
            if position % 50 == 0:
                row_residues = row_residues + "\n"
                row_features = row_features + "\n"
            continue

        # residue not adaptive
        if float(values[2]) < 0.70 and values[3] != 0:
            residue = values[0]
            row_features = row_features + " "
            row_residues = row_residues + residue
            if position % 50 == 0:
                row_residues = row_residues + "\n"
                row_features = row_features + "\n"
            
        # residue with mild probability of adaptive evolution
        if 0.70 <= float(values[2]) < 0.90:
            residue = values[0]
            row_features = row_features + "."
            row_residues = row_residues + residue
            if position % 50 == 0:
                row_residues = row_residues + "\n"
                row_features = row_features + "\n"
                
        # residue with strong probability of adaptive evolution
        if 0.90 <= float(values[2]):
            residue = values[0]
            row_features = row_features + ":"
            row_residues = row_residues + residue
            if position % 50 == 0:
                row_residues = row_residues + "\n"
                row_features = row_features + "\n"

    positions = row_positions.split("\n")
    features = row_features.split("\n")
    residues = row_residues.split("\n")
    
    if len(positions) == len(features) == len(residues):
    
        message = ("\n\n.........................................." 
                   "\n\nReference sequence for %s with adaptive sites:"
                   "\n\n . means pr >= 0.70 \n : means pr >= 0.90 \n - means missing from PAML analysis\n\n") % gene
        print(message)
        logging.info(message)
        
        for row in range(len(positions)):
            message = positions[row]
            print(message)
            logging.info(message)
            
            message = features[row]
            print(message)
            logging.info(message)
            
            message = residues[row]
            print(message)
            logging.info(message)
            
    else:
        message = "problem with unequal rows"
        print(message)
        print(positions)
        print(features)
        print(residues)
        logging.info(message)
        logging.info(positions)
        logging.info(features)
        logging.info(residues)


def get_ref_protein_seq(wdir, gene, ref_species):
    """Gets the protein sequenced used for blast"""

    # open according cds fasta file
    with open(wdir + "Blast_input/" + gene + "_" + ref_species + "_protein.fasta", "r") as f:
        sequence = ""
        cds = f.readlines()
        for line in cds[1:]:
            sequence = sequence + line.rstrip("\n")
    return sequence
    

def get_ref_and_final_seqs(wdir, gene, result_path, ref_species):
    """Recovers aa seq of reference species from input and a final one after Gblocks"""
    
    gene_path = result_path + "/" + gene

    # get record object for the ref protein sequence
    ref_sequence = get_ref_protein_seq(wdir, gene, ref_species)
    ref_sequence_Seq_object = Seq(ref_sequence)
    ref_sequence_record = SeqRecord(ref_sequence_Seq_object)
    
    # get record object for the final protein sequence
    pattern = "translated.fasta"
    all_files = os.listdir(gene_path)
    
    for file in all_files:
        if pattern in file and "reduced" not in file and not file.startswith("."):
            post_Gblocks_translated_file_path = gene_path + "/" + file
    
    alignment = AlignIO.read(post_Gblocks_translated_file_path, "fasta")
    all_seqs = [fasta_reader.read_fasta_record(record.format("fasta")) for record in alignment]
    final_sequence = all_seqs[0][1]
    final_sequence_Seq_object = Seq(final_sequence)
    final_sequence_record = SeqRecord(final_sequence_Seq_object)
    
    return ref_sequence_record, final_sequence_record


def organise_ref_and_final_seqs(ref_sequence_record, final_sequence_record):
    """Organises residues in aa seq of reference species from input and post Gblocks into an easy to search dict"""

    # aa seq from input
    ref_species_dict = {}
    total_length = 0
    for index, aa in enumerate(ref_sequence_record.seq, start=1):
        if aa == "*":
            break
        ref_species_dict[index] = aa
        total_length += 1

    # aa seq post Gblocks
    final_ref_species_dict = {}
    final_length = 0
    for index, aa in enumerate(final_sequence_record.seq, start=1):
        if aa == "*":
            break
        final_ref_species_dict[index] = aa
        final_length += 1

    return ref_species_dict, final_ref_species_dict, final_length


def map_ref_and_final_residues(ref_sequence_record, final_sequence_record):
    """Maps residues between the aa seq of reference species and post Gblocks"""
    
    # perform pairwise alignment (multiple alignments of the same sequences)
    # x - matches = 1, mismatches = 0; s - gap penalty (second parameter x would be no gap penalties)
    aln = pairwise2.align.globalxs(final_sequence_record.seq, ref_sequence_record.seq, open=-0.5, extend=-0.1)
    
    # make a dict storing the alignments
    mapped_ref_and_final_residues_dict = {}
    for i in range(1, len(aln[0][0]) + 1):  # added "+1" on 09_04_2021
        mapped_ref_and_final_residues_dict[i] = []

    # fill the dict with paired positions (need to subtract 1 cose aln starts at 0)
    for position in mapped_ref_and_final_residues_dict:
        mapped_ref_and_final_residues_dict[position] = [aln[0][0][position-1], aln[0][1][position-1]]
    
    # flag any non-synonymous substitutions that indicate frameshifts
    frameshift_positions = {}
    translated_frameshift = False
    for position, pair in mapped_ref_and_final_residues_dict.items():
        if pair[0] != pair[1] and pair[0] != "-" and pair[1] != "-":
            # sequence index starts at 1 so need to "+1" from the python indexing
            frameshift_positions[position + 1] = pair
            translated_frameshift = True
            print("...WARNING... : position - %s (%s) between final protein ref seq differs from blast input "
                  "-> frameshift?" % (position + 1, (pair[0], pair[1])))

    return mapped_ref_and_final_residues_dict


def match_adaptive_sites_to_ref(final_ref_species_dict, mapped_ref_and_final_residues_dict, adaptive_sites_dict,
                                omega_dict):
    """Matches sites under positive selection from seq post Gblocks onto the one used for blast input (ref species)"""
    
    # overlay the final residue dictionary with probability for positive selection
    matched_adaptive_sites_final = {}
    for site, aa in final_ref_species_dict.items():
        # mark this site as potentially positively selected
        if site in adaptive_sites_dict:
            matched_adaptive_sites_final[site] = adaptive_sites_dict[site]
        # mark this site as not positively selected
        else:
            matched_adaptive_sites_final[site] = [aa, "0.00"]
    
    # adjust the final residue dictionary to include missing sites
    adjusted_mapped_ref_and_final_residue_dict = {}
    for i in range(1, len(mapped_ref_and_final_residues_dict) + 1):
        adjusted_mapped_ref_and_final_residue_dict[i] = ["", "", 0]
    
    a = mapped_ref_and_final_residues_dict
    b = adjusted_mapped_ref_and_final_residue_dict
    for site, aa in a.items():
        # special case when there is no previous site/key
        if site == 1 and aa[0] == "-":
            b[site] = [aa[0], aa[1], 1]
            continue
        elif site == 1 and aa[0] != "-":
            b[site] = [aa[0], aa[1], 0]
        # collect dashes and add them to tally per site
        elif site != 1 and aa[0] == "-":
            b[site] = [aa[0], aa[1], b[site-1][2]+1]
        # propagate the tally from previous site
        elif site != 1 and aa[0] != "-":
            b[site] = [aa[0], aa[1], b[site-1][2]]
        # safety print -> REMOVE AFTER TESTING
        else:
            print("else")
    
    # map omegas and probabilities to the ref residues
    c = matched_adaptive_sites_final
    w = omega_dict
    matched_adaptive_sites_ref = {}
    for site, aa in a.items():
        if aa[0] == aa[1]:
            key = [k for k, v in b.items() if k == site][0]
            probability = c[key-b[site][2]][1]
            omega = w[key-b[site][2]]
            matched_adaptive_sites_ref[site] = [aa[0], omega, probability]
        else:
            matched_adaptive_sites_ref[site] = [aa[1], "0.000", "0.000"]

    return matched_adaptive_sites_ref


def get_omegas(gene, result_path, final_length):
    """Gets Dn/Ds ratio (omega) for each site in aa seq of reference species"""
    
    PAML_output_file_path = result_path + gene + "/PAML_" + gene
    
    PAML_result_dict = {}
    for i in range(1, final_length+1):
        PAML_result_dict[i] = ""

    # collect all values for each residue from the BEB of M8 model
    with open(PAML_output_file_path + "/rst", "r") as f:
        
        start_search = False
        start_recording = False
        line_number = 0
        # assumes that max protein length == 9999aa
        space = " " * (4 - len(str(final_length)))
        file = f.readlines()
        
        l = 0
        for line in file:
            l += 1
            
            if line_number == final_length:
                break
            
            if start_recording is False and line.startswith("Model 8"):
                start_search = True
                continue
            
            if start_search is True and line.startswith("Bayes"):
                start_recording = True
                continue
            
            if start_recording is True and line.startswith("   1 "):
                line_number += 1
                PAML_result_dict[line_number] = line.split(" ")
                continue
            
            if start_recording is True and PAML_result_dict[1] != "":
                line_number += 1
                PAML_result_dict[line_number] = line.split(" ")
                continue
            
            if start_recording is True and line.startswith(space + str(final_length)):
                line_number += 1
                PAML_result_dict[line_number] = line.split(" ")
                start_recording = False
    
    # collect posterior mean omegas for each residue
    omega_dict = {}
    for residue, values in PAML_result_dict.items():
        # remove empty strings which can change number of elements in values
        values = [element for element in values if element != ""]
        # record value only for recurrently changing sites where omega is likely > 1 (not all > 1 are in M8 class 11 !!)
        if int((values[-4]).lstrip("(").rstrip(")")) == 11:
            omega_dict[residue] = float(values[-3])
        else:
            omega_dict[residue] = 0.000
        
    return omega_dict


def get_adaptive_sites(result_path, gene):
    """Reads the PAML output files and finds which sites are likely under positive selection (BEB)"""
    
    PAML_output_file_path = result_path + "/" + gene + "/PAML_" + gene
    with open(PAML_output_file_path + "/output_PAML", "r") as f:
        
        adaptive_sites_dict = {}
        start_search1 = False
        start_search2 = False
        start_recording = False
        newlines = 0
        file = f.readlines()
        
        for line in file:

            if start_recording is True and re.search(r"^\s*\d", line):
                site_number = int(re.match(r"^\s*\d{1,4}", line).group(0).replace(" ",""))
                residue = re.match(r"^\s*\d{1,4}\s[A-Z]", line).group(0)[-1]
                probability = re.match(r"^\s*\d{1,4}\s[A-Z]\s*\d\.\d{3}", line).group(0)[-5:]
                adaptive_sites_dict[site_number] = [residue, probability]
                continue
                
            if start_recording is True and line.startswith("\n"):
                newlines += 1
                
            if newlines == 2:
                break

            if start_recording is False and line.startswith("Model 8"):
                start_search1 = True
                continue
            
            if start_search1 is True and line.startswith("Bayes"):
                start_search2 = True
                continue
            
            # 12 spaces in front of "Pr"
            if start_search2 is True and line.startswith("            Pr"):
                start_recording = True
                continue
        
    return adaptive_sites_dict


def make_graphs(wdir, ref_species, final_dict_to_plot, result_path, gene, nr_of_species_total,
                genes_under_positive_selection):
    """Draws a graph of PAML analysis : omegas, all posterior probabilities and highly likely sites
    under positive selection"""

    sites = []
    residues = []
    omegas = []
    probabilities = []
    present = []
    # these will determine omega graph y axis
    roof = 2
    floor = 1

    for site, features in final_dict_to_plot.items():
        sites.append(site)
        residues.append(features[0])
        
        if features[3] == 1:
            omegas.append(float(features[1]))
            if float(features[1]) > roof:
                roof = math.ceil(features[1])
            elif 0.5 <= float(features[1]) <= floor:
                floor = 0.5
            elif 0 < float(features[1]) < 0.5:
                floor = 0
            probabilities.append(float(features[2]))
            
        # mark missing values as impossibly high omegas and probabilities
        else:
            omegas.append(float(20.00))
            probabilities.append(float(20.00))
    
        present.append(features[3])

    # reset probabilities for genes unlikely under positive selection
    if gene not in genes_under_positive_selection:
        for i in range(len(probabilities)):
            # mark missing residues
            if present[i] == 0:
                probabilities[i] = 20
            # dont show any probabilities for present residues
            else:
                probabilities[i] = 0

    # set the backend to a non-interactive one -> does not create and destroys GUI windows
    use('agg')
    plt.figure()

    if ref_species == "Mm" or ref_species == "Rn":
        clade = "Rodents"
    elif ref_species == "Hs":
        clade = "Primates"
    elif ref_species == "Fc" or ref_species == "Cf":
        clade = "Carnivora"
    elif ref_species == "Gg":
        clade = "Phasanidae"

    # plot recurrently changing sites (put it 11 bin of M8 model with postmean omega > 1.0)
    plt.subplot(311, title="PAML analysis - %s (%s - %s species)" % (gene, clade, nr_of_species_total))
    plt.ylabel("Posterior mean\n omega")
    plt.axis([1.0, sites[-1], 0, roof + 1.1])
    plt.ylim(1.0, roof + 1.1)
    plt.yticks(np.arange(1.0, roof + 1.1, 1.0))
    # mark the missing values for the plot
    clrs1 = ["black" if s == 1 else "gainsboro" for s in present]
    plt.bar(sites, omegas, color=clrs1)

    # plot posterior probabilities from BEB analysis (pr >= 0.70)
    plt.subplot(312)
    plt.ylabel("Prob. positive\n selection")
    plt.axis([0, sites[-1], 0.7, 1.0])
    plt.ylim(0.7, 1.0)
    plt.yticks(np.arange(0.7, 1.0, 0.1))
    # mark the missing values for the plot
    clrs2 = ["lightblue" if s == 1 else "gainsboro" for s in present]
    plt.bar(sites, probabilities, color=clrs2)
    
    # plot residues with highest posterior probabilities from BEB analysis (pr >= 0.90)
    plt.subplot(313)
    plt.xlabel("Codons (mapped to CDS used as reference)")
    plt.ylabel("High prob. \n positive selection")
    plt.axis([0, sites[-1], 0.9, 1.0])
    plt.yticks(np.arange(0.9, 1.01, 0.05))
    # mark the missing values for the plot
    clrs3 = ["magenta" if s == 1 else "gainsboro" for s in present]
    plt.bar(sites, probabilities, color=clrs3)
    plt.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
    
    figure_name = gene + "_PAML_analysis"
    plt.savefig(figure_name + ".tif", dpi=300, bbox_inches="tight")
    
    shutil.move(wdir + figure_name + ".tif", result_path + figure_name + ".tif")


    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
        
